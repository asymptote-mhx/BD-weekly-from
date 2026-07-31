from __future__ import annotations

from dataclasses import asdict, fields
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from market_workbench.config import (
    CHAIN_PEOPLE_HEADERS,
    COMPETITOR_HEADERS,
    DETAIL_HEADERS,
    MARKETING_EVENT_HEADERS,
    ORG_HEADERS,
    PEOPLE_HEADERS,
    PROGRESS_HEADERS,
    PROJECT_HEADERS,
    PUBLIC_PROGRESS,
    SENSITIVE_HEADERS,
    SHEET_WEEKLY_IMPORTS,
    SHEET_CHAIN_PEOPLE,
    SHEET_COMPETITORS,
    SHEET_ORGS,
    SHEET_PEOPLE,
    SHEET_PROGRESS,
    SHEET_PROJECTS,
    SHEET_SENSITIVE,
    SHEET_DETAILS,
    SHEET_MARKETING_EVENTS,
    WEEKLY_IMPORT_HEADERS,
)
from market_workbench.models import Project, SensitiveBusiness


PROJECT_FIELDS = {field.name for field in fields(Project)}

WORKBOOK_SHEETS = (
    (SHEET_PROJECTS, PROJECT_HEADERS),
    (SHEET_ORGS, ORG_HEADERS),
    (SHEET_PEOPLE, PEOPLE_HEADERS),
    (SHEET_PROGRESS, PROGRESS_HEADERS),
    (SHEET_SENSITIVE, SENSITIVE_HEADERS),
    (SHEET_DETAILS, DETAIL_HEADERS),
    (SHEET_CHAIN_PEOPLE, CHAIN_PEOPLE_HEADERS),
    (SHEET_MARKETING_EVENTS, MARKETING_EVENT_HEADERS),
    (SHEET_COMPETITORS, COMPETITOR_HEADERS),
    (SHEET_WEEKLY_IMPORTS, WEEKLY_IMPORT_HEADERS),
)

PROJECT_FIELD_NAMES = (
    "project_id",
    "name",
    "region",
    "owner_org",
    "partner_org",
    "owner_type",
    "public_progress",
    "detail_stage",
    "next_node_time",
    "next_work",
    "needs_technical_input",
    "technical_type",
    "internal_owner",
    "status_note",
    "updated_at",
    "confirmation_status",
    "detailed_address",
    "land_area",
    "building_area",
    "construction_scale",
    "total_investment",
    "estimated_contract_amount",
    "construction_content",
    "planning_design_scope",
    "parent_project_id",
    "parent_project_name",
    "relationship_reason",
    "record_status",
    "priority",
)

PROJECT_FIELD_TO_HEADER = dict(zip(PROJECT_FIELD_NAMES, PROJECT_HEADERS))

PROJECT_HEADER_TO_FIELD = {header: field for field, header in PROJECT_FIELD_TO_HEADER.items()}

SENSITIVE_FIELD_TO_HEADER = {
    "project_id": "project_id",
    "estimated_design_fee": "预计设计费用",
    "quote_range": "报价区间",
    "business_cost": "商务成本",
    "competition": "竞争格局",
    "advantage": "切入优势",
    "risks": "风险点",
    "note": "商务备注",
    "updated_at": "最近更新时间",
}


class ExcelStore:
    def __init__(self, path: Path):
        self.path = Path(path)

    def ensure_workbook(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.exists():
            workbook = load_workbook(self.path)
        else:
            workbook = Workbook()

        changed = False
        for sheet_name, headers in WORKBOOK_SHEETS:
            changed = self._ensure_sheet(workbook, sheet_name, headers) or changed
            if sheet_name == SHEET_PROJECTS:
                changed = self._repair_shifted_project_rows(workbook[sheet_name]) or changed

        for sheet_name in list(workbook.sheetnames):
            if sheet_name == "Sheet" and len(workbook.sheetnames) > 1:
                del workbook[sheet_name]
                changed = True

        if changed:
            workbook.save(self.path)
        workbook.close()

    def list_projects(self) -> list[dict[str, str]]:
        workbook = self._load()
        try:
            sheet = workbook[SHEET_PROJECTS]
            return self._list_rows(sheet, PROJECT_HEADERS)
        finally:
            workbook.close()

    def list_weekly_imports(self) -> list[dict[str, str]]:
        workbook = self._load()
        try:
            sheet = workbook[SHEET_WEEKLY_IMPORTS]
            return self._list_rows(sheet, WEEKLY_IMPORT_HEADERS)
        finally:
            workbook.close()

    def list_progress_records(self) -> list[dict[str, str]]:
        workbook = self._load()
        try:
            sheet = workbook[SHEET_PROGRESS]
            return self._list_rows(sheet, PROGRESS_HEADERS)
        finally:
            workbook.close()

    def record_weekly_import(
        self,
        source_file: str,
        report_date: str,
        project_count: int,
        note: str = "",
    ) -> dict[str, str]:
        row = {
            "source_file": source_file,
            "报告日期": report_date,
            "导入时间": self._now(),
            "项目数量": str(project_count),
            "备注": note,
        }
        workbook = self._load()
        try:
            sheet = workbook[SHEET_WEEKLY_IMPORTS]
            self._upsert_row_by_key(sheet, WEEKLY_IMPORT_HEADERS, row, "source_file")
            workbook.save(self.path)
            return row
        finally:
            workbook.close()

    def append_progress_record(self, row: dict[str, str]) -> dict[str, str]:
        values = {header: "" if row.get(header) is None else str(row.get(header, "")) for header in PROGRESS_HEADERS}
        workbook = self._load()
        try:
            sheet = workbook[SHEET_PROGRESS]
            self._append_row(sheet, PROGRESS_HEADERS, values)
            workbook.save(self.path)
            return values
        finally:
            workbook.close()

    def delete_unconfirmed_projects(self) -> int:
        workbook = self._load()
        try:
            sheet = workbook[SHEET_PROJECTS]
            status_column = PROJECT_HEADERS.index("数据确认状态") + 1
            deleted = 0

            for row_number in range(sheet.max_row, 1, -1):
                status = str(sheet.cell(row=row_number, column=status_column).value or "")
                if status != "已确认":
                    sheet.delete_rows(row_number)
                    deleted += 1

            if deleted:
                workbook.save(self.path)
            return deleted
        finally:
            workbook.close()

    def delete_project(self, project_id: str) -> dict[str, str]:
        workbook = self._load()
        try:
            sheet = workbook[SHEET_PROJECTS]
            deleted = self._delete_project_row(sheet, project_id)
            if deleted is None:
                raise KeyError(project_id)
            self._delete_related_project_rows(workbook, project_id)
            workbook.save(self.path)
            return deleted
        finally:
            workbook.close()

    def delete_projects_by_record_status(self, status: str) -> int:
        workbook = self._load()
        try:
            sheet = workbook[SHEET_PROJECTS]
            status_column = PROJECT_HEADERS.index("记录状态") + 1
            deleted = 0

            for row_number in range(sheet.max_row, 1, -1):
                row_status = str(sheet.cell(row=row_number, column=status_column).value or "")
                if row_status == status:
                    project_id = str(
                        sheet.cell(
                            row=row_number,
                            column=PROJECT_HEADERS.index("project_id") + 1,
                        ).value
                        or ""
                    )
                    sheet.delete_rows(row_number)
                    if project_id:
                        self._delete_related_project_rows(workbook, project_id)
                    deleted += 1

            if deleted:
                workbook.save(self.path)
            return deleted
        finally:
            workbook.close()

    def upsert_project(self, project: Project) -> dict[str, str]:
        workbook = self._load()
        try:
            sheet = workbook[SHEET_PROJECTS]
            row = self._project_to_row(project)
            if not row.get("最近更新时间"):
                row["最近更新时间"] = self._now()
            self._upsert_row(sheet, PROJECT_HEADERS, row)
            workbook.save(self.path)
            return row
        finally:
            workbook.close()

    def upsert_project_dict(self, payload: dict[str, str]) -> dict[str, str]:
        values = {}
        for key, value in payload.items():
            field = PROJECT_HEADER_TO_FIELD.get(key, key)
            if field in PROJECT_FIELDS:
                values[field] = "" if value is None else str(value)

        values.setdefault("confirmation_status", "已确认")
        values.setdefault("record_status", "正常")

        return self.upsert_project(Project(**values))

    def get_project(self, project_id: str) -> dict[str, str] | None:
        for row in self.list_projects():
            if row.get("project_id") == project_id:
                return row
        return None

    def set_project_record_status(self, project_id: str, status: str) -> dict[str, str]:
        row = self.get_project(project_id)
        if row is None:
            raise KeyError(project_id)
        row["记录状态"] = status
        return self.upsert_project_dict(row)

    def merge_project(
        self, project_id: str, main_project_id: str, reason: str = ""
    ) -> dict[str, str]:
        project = self.get_project(project_id)
        main_project = self.get_project(main_project_id)
        if project is None:
            raise KeyError(project_id)
        if main_project is None:
            raise KeyError(main_project_id)
        if project_id == main_project_id:
            raise ValueError("Cannot merge a project into itself")

        project["记录状态"] = "已合并"
        project["主项目ID"] = main_project_id
        project["主项目名称"] = main_project.get("项目名称", "")
        project["关联原因"] = reason
        written = self.upsert_project_dict(project)

        source_note = (
            f"已合并项目：{project.get('项目名称', '')}"
            f"；原因：{reason or '-'}"
            f"；原下一步：{project.get('下一步工作', '') or '-'}"
        )
        existing_note = main_project.get("状态备注", "").strip()
        main_project["状态备注"] = (
            f"{existing_note}\n{source_note}" if existing_note else source_note
        )
        self.upsert_project_dict(main_project)
        return written

    def upsert_sensitive(self, sensitive: SensitiveBusiness) -> dict[str, str]:
        workbook = self._load()
        try:
            sheet = workbook[SHEET_SENSITIVE]
            row = self._sensitive_to_row(sensitive)
            if not row.get("最近更新时间"):
                row["最近更新时间"] = self._now()
            self._upsert_row(sheet, SENSITIVE_HEADERS, row)
            workbook.save(self.path)
            return row
        finally:
            workbook.close()

    def get_sensitive(self, project_id: str) -> dict[str, str]:
        workbook = self._load()
        try:
            sheet = workbook[SHEET_SENSITIVE]
            for row in self._list_rows(sheet, SENSITIVE_HEADERS):
                if row.get("project_id") == project_id:
                    return row
            return {"project_id": project_id}
        finally:
            workbook.close()

    def get_project_detail(self, project_id: str) -> dict[str, str]:
        workbook = self._load()
        try:
            sheet = workbook[SHEET_DETAILS]
            for row in self._list_rows(sheet, DETAIL_HEADERS):
                if row.get("project_id") == project_id:
                    return row
            return {header: "" for header in DETAIL_HEADERS} | {"project_id": project_id}
        finally:
            workbook.close()

    def upsert_project_detail(self, payload: dict[str, str]) -> dict[str, str]:
        workbook = self._load()
        try:
            sheet = workbook[SHEET_DETAILS]
            row = {
                header: "" if payload.get(header) is None else str(payload.get(header, ""))
                for header in DETAIL_HEADERS
            }
            if not row.get("project_id"):
                raise ValueError("project_id is required")
            if not row.get("最近更新时间"):
                row["最近更新时间"] = self._now()
            self._upsert_row(sheet, DETAIL_HEADERS, row)
            workbook.save(self.path)
            return row
        finally:
            workbook.close()

    def get_project_structured_detail(self, project_id: str) -> dict[str, list[dict[str, str]]]:
        workbook = self._load()
        try:
            return {
                "chain_people": self._list_project_rows(
                    workbook[SHEET_CHAIN_PEOPLE], CHAIN_PEOPLE_HEADERS, project_id
                ),
                "marketing_events": self._list_project_rows(
                    workbook[SHEET_MARKETING_EVENTS], MARKETING_EVENT_HEADERS, project_id
                ),
                "competitors": self._list_project_rows(
                    workbook[SHEET_COMPETITORS], COMPETITOR_HEADERS, project_id
                ),
            }
        finally:
            workbook.close()

    def replace_project_structured_detail(
        self, project_id: str, payload: dict[str, list[dict[str, str]]]
    ) -> dict[str, list[dict[str, str]]]:
        workbook = self._load()
        try:
            chain_people = self._replace_project_rows(
                workbook[SHEET_CHAIN_PEOPLE],
                CHAIN_PEOPLE_HEADERS,
                project_id,
                payload.get("chain_people", []),
                item_prefix="person",
            )
            marketing_events = self._replace_project_rows(
                workbook[SHEET_MARKETING_EVENTS],
                MARKETING_EVENT_HEADERS,
                project_id,
                payload.get("marketing_events", []),
                item_prefix="event",
            )
            competitors = self._replace_project_rows(
                workbook[SHEET_COMPETITORS],
                COMPETITOR_HEADERS,
                project_id,
                payload.get("competitors", []),
                item_prefix="competitor",
            )
            workbook.save(self.path)
            return {
                "chain_people": chain_people,
                "marketing_events": marketing_events,
                "competitors": competitors,
            }
        finally:
            workbook.close()

    def _load(self):
        self.ensure_workbook()
        return load_workbook(self.path)

    @staticmethod
    def _ensure_sheet(workbook, sheet_name: str, headers: list[str]) -> bool:
        changed = False
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        elif workbook.active.title == "Sheet" and workbook.active.max_row == 1:
            sheet = workbook.active
            sheet.title = sheet_name
            changed = True
        else:
            sheet = workbook.create_sheet(sheet_name)
            changed = True

        existing_headers = [
            sheet.cell(row=1, column=column).value
            for column in range(1, max(sheet.max_column, len(headers)) + 1)
        ]
        if existing_headers[: len(headers)] != headers and any(
            header in headers for header in existing_headers
        ):
            rows = []
            for row_number in range(2, sheet.max_row + 1):
                row = {
                    header: sheet.cell(row=row_number, column=column).value
                    for column, header in enumerate(existing_headers, start=1)
                    if header
                }
                if any(value not in (None, "") for value in row.values()):
                    rows.append(row)

            for row_number in range(1, sheet.max_row + 1):
                for column in range(1, sheet.max_column + 1):
                    sheet.cell(row=row_number, column=column, value=None)

            for column, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=column, value=header)
            for row_number, row in enumerate(rows, start=2):
                for column, header in enumerate(headers, start=1):
                    sheet.cell(row=row_number, column=column, value=row.get(header, ""))
            return True

        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=column)
            if cell.value != header:
                cell.value = header
                changed = True
        return changed

    @staticmethod
    def _repair_shifted_project_rows(sheet: Worksheet) -> bool:
        index = {header: column for column, header in enumerate(PROJECT_HEADERS, start=1)}
        confirmation_values = {"草稿", "已确认"}
        record_status_values = {"正常", "已归档", "已合并", "已删除"}
        changed = False

        for row_number in range(2, sheet.max_row + 1):
            owner_type = str(sheet.cell(row=row_number, column=index["业主类型"]).value or "")
            updated_at = str(sheet.cell(row=row_number, column=index["最近更新时间"]).value or "")
            relation_reason = str(sheet.cell(row=row_number, column=index["关联原因"]).value or "")

            if not (
                owner_type in PUBLIC_PROGRESS
                and updated_at in confirmation_values
                and relation_reason in record_status_values
            ):
                continue

            for column in range(index["关联原因"], index["业主类型"] - 1, -1):
                sheet.cell(
                    row=row_number,
                    column=column + 1,
                    value=sheet.cell(row=row_number, column=column).value,
                )
            sheet.cell(row=row_number, column=index["业主类型"], value="")
            changed = True

        return changed

    @staticmethod
    def _list_rows(sheet: Worksheet, headers: list[str]) -> list[dict[str, str]]:
        rows = []
        for values in sheet.iter_rows(min_row=2, max_col=len(headers), values_only=True):
            if all(value in (None, "") for value in values):
                continue
            rows.append(
                {
                    header: "" if value is None else str(value)
                    for header, value in zip(headers, values)
                }
            )
        return rows

    @classmethod
    def _list_project_rows(
        cls, sheet: Worksheet, headers: list[str], project_id: str
    ) -> list[dict[str, str]]:
        rows = [
            row
            for row in cls._list_rows(sheet, headers)
            if row.get("project_id") == project_id
        ]
        return sorted(rows, key=lambda row: cls._row_sort_index(row))

    @staticmethod
    def _row_sort_index(row: dict[str, str]) -> tuple[int, str]:
        try:
            order = int(row.get("排序", "") or "0")
        except ValueError:
            order = 0
        return order, row.get("item_id", "")

    def _replace_project_rows(
        self,
        sheet: Worksheet,
        headers: list[str],
        project_id: str,
        rows: list[dict[str, str]],
        item_prefix: str,
    ) -> list[dict[str, str]]:
        project_column = headers.index("project_id") + 1
        for row_number in range(sheet.max_row, 1, -1):
            if str(sheet.cell(row=row_number, column=project_column).value or "") == project_id:
                sheet.delete_rows(row_number)

        written_rows = []
        for index, source in enumerate(rows, start=1):
            if not self._has_meaningful_detail_values(source, headers):
                continue
            row = {
                header: "" if source.get(header) is None else str(source.get(header, ""))
                for header in headers
            }
            row["project_id"] = project_id
            row["item_id"] = row.get("item_id") or f"{project_id}-{item_prefix}-{index}"
            row["排序"] = row.get("排序") or str(index)
            row["最近更新时间"] = row.get("最近更新时间") or self._now()
            self._append_row(sheet, headers, row)
            written_rows.append(row)
        return written_rows

    @staticmethod
    def _delete_rows_by_project_id(sheet: Worksheet, headers: list[str], project_id: str) -> None:
        project_column = headers.index("project_id") + 1
        for row_number in range(sheet.max_row, 1, -1):
            if str(sheet.cell(row=row_number, column=project_column).value or "") == project_id:
                sheet.delete_rows(row_number)

    def _delete_related_project_rows(self, workbook, project_id: str) -> None:
        for sheet_name, headers in (
            (SHEET_SENSITIVE, SENSITIVE_HEADERS),
            (SHEET_DETAILS, DETAIL_HEADERS),
            (SHEET_CHAIN_PEOPLE, CHAIN_PEOPLE_HEADERS),
            (SHEET_MARKETING_EVENTS, MARKETING_EVENT_HEADERS),
            (SHEET_COMPETITORS, COMPETITOR_HEADERS),
        ):
            self._delete_rows_by_project_id(workbook[sheet_name], headers, project_id)

    @staticmethod
    def _has_meaningful_detail_values(row: dict[str, str], headers: list[str]) -> bool:
        ignored = {"item_id", "project_id", "排序", "最近更新时间"}
        return any(str(row.get(header, "")).strip() for header in headers if header not in ignored)

    @staticmethod
    def _upsert_row(sheet: Worksheet, headers: list[str], row: dict[str, str]) -> None:
        target_row = None
        for row_number in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row_number, column=1).value or "") == row["project_id"]:
                target_row = row_number
                break

        if target_row is None:
            target_row = sheet.max_row + 1

        for column, header in enumerate(headers, start=1):
            sheet.cell(row=target_row, column=column, value=row.get(header, ""))

    @staticmethod
    def _upsert_row_by_key(
        sheet: Worksheet, headers: list[str], row: dict[str, str], key_header: str
    ) -> None:
        key_column = headers.index(key_header) + 1
        key_value = str(row.get(key_header, ""))
        target_row = None
        for row_number in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row_number, column=key_column).value or "") == key_value:
                target_row = row_number
                break

        if target_row is None:
            target_row = sheet.max_row + 1

        for column, header in enumerate(headers, start=1):
            sheet.cell(row=target_row, column=column, value=row.get(header, ""))

    @staticmethod
    def _append_row(sheet: Worksheet, headers: list[str], row: dict[str, str]) -> None:
        target_row = sheet.max_row + 1
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=target_row, column=column, value=row.get(header, ""))

    @staticmethod
    def _delete_project_row(sheet: Worksheet, project_id: str) -> dict[str, str] | None:
        project_id_column = PROJECT_HEADERS.index("project_id") + 1
        for row_number in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row_number, column=project_id_column).value or "") == project_id:
                row = {
                    header: str(sheet.cell(row=row_number, column=column).value or "")
                    for column, header in enumerate(PROJECT_HEADERS, start=1)
                }
                sheet.delete_rows(row_number)
                return row
        return None

    @staticmethod
    def _project_to_row(project: Project) -> dict[str, str]:
        source = asdict(project)
        return {
            header: "" if source[field] is None else str(source[field])
            for field, header in PROJECT_FIELD_TO_HEADER.items()
        }

    @staticmethod
    def _sensitive_to_row(sensitive: SensitiveBusiness) -> dict[str, str]:
        source = asdict(sensitive)
        return {
            header: "" if source[field] is None else str(source[field])
            for field, header in SENSITIVE_FIELD_TO_HEADER.items()
        }

    @staticmethod
    def _now() -> str:
        return datetime.now().isoformat(timespec="seconds")
